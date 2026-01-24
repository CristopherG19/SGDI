"""
SGDI - Servicio Procesador de Excel
====================================

Procesa archivos Excel para generar documentos con QR:
- Lee datos de una hoja de datos
- Genera QRs en batch
- Llena plantilla con cada registro
- Exporta a PDF
- Opcionalmente imprime
"""

import os
import time
from pathlib import Path
from typing import Dict, List, Callable, Optional, Tuple, Any
import openpyxl

from core.utils.logger import get_logger, log_operation
from core.database.simple_db import get_db
from modules.qr_suite.services.qr_generator import QRGenerator

log = get_logger(__name__)


class ExcelProcessor:
    """Procesador de archivos Excel con generación de QR."""
    
    def __init__(self):
        """Inicializa el procesador."""
        self.db = get_db()
        self.qr_generator = QRGenerator()
        self.stop_requested = False
    
    def stop(self):
        """Detiene el procesamiento."""
        self.stop_requested = True
        log.info("Deteniendo procesamiento de Excel...")
    
    def get_sheet_names(self, excel_path: str | Path) -> List[str]:
        """
        Obtiene nombres de hojas del Excel.
        
        Args:
            excel_path: Ruta al archivo Excel
            
        Returns:
            Lista de nombres de hojas
        """
        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            return sheets
        except Exception as e:
            log.error(f"Error leyendo hojas: {e}")
            return []
    
    def read_data_sheet(self, 
                       excel_path: str | Path,
                       sheet_name: str = "DATA",
                       key_column: int = 17) -> Tuple[List[tuple], List[str]]:
        """
        Lee datos de la hoja de datos.
        
        Args:
            excel_path: Ruta al Excel
            sheet_name: Nombre de la hoja de datos
            key_column: Columna clave para QR (0-indexed, default 17 = columna R)
            
        Returns:
            Tupla (lista_filas, lista_claves_qr)
        """
        try:
            log.info(f"Leyendo datos de {sheet_name}...")
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            ws = wb[sheet_name]
            
            rows = []
            keys = []
            
            # Saltar encabezados (fila 1)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue  # Saltar filas vacías
                    
                rows.append(row)
                
                # Obtener clave para QR
                key = str(row[key_column]) if len(row) > key_column and row[key_column] else "SIN_CLAVE"
                keys.append(key)
            
            wb.close()
            log.info(f"Leídos {len(rows)} registros")
            return rows, keys
            
        except Exception as e:
            log.error(f"Error leyendo datos: {e}")
            return [], []
    
    def generate_qr_batch(self, 
                         keys: List[str],
                         output_folder: str | Path) -> Dict[int, str]:
        """
        Genera QRs en batch para las claves.
        
        Args:
            keys: Lista de contenidos para QR
            output_folder: Carpeta de salida
            
        Returns:
            Diccionario {índice: ruta_qr}
        """
        output_folder = Path(output_folder)
        output_folder.mkdir(parents=True, exist_ok=True)
        
        qr_paths = {}
        
        log.info(f"Generando {len(keys)} códigos QR...")
        
        for idx, key in enumerate(keys):
            if self.stop_requested:
                break
                
            try:
                qr_path = output_folder / f"qr_{idx}_{key}.png"
                success, result = self.qr_generator.generate_qr(
                    content=key,
                    output_path=str(qr_path),
                    size=200,
                    error_correction='M'
                )
                
                if success:
                    qr_paths[idx] = str(qr_path)
                else:
                    log.warning(f"No se generó QR para {key}: {result}")
                    
            except Exception as e:
                log.error(f"Error generando QR {idx}: {e}")
        
        log.info(f"Generados {len(qr_paths)} QRs")
        return qr_paths
    
    def process_with_com(self,
                        config: Dict[str, Any],
                        progress_callback: Optional[Callable] = None) -> Dict[str, int]:
        """
        Procesa Excel usando COM de Windows (requiere Excel instalado).
        
        Args:
            config: Configuración del proceso:
                - archivo_excel: Ruta al archivo
                - hoja_plantilla: Nombre de la hoja plantilla
                - hoja_datos: Nombre de la hoja de datos
                - columna_clave: Columna para QR (0-indexed)
                - generar_pdf: Si generar PDFs
                - carpeta_pdf: Carpeta de salida PDFs
                - imprimir: Si imprimir
                - impresora: Nombre de impresora
                - copias: Número de copias
            progress_callback: Función(idx, total, actual)
            
        Returns:
            Estadísticas del proceso
        """
        import pythoncom
        import win32com.client
        
        self.stop_requested = False
        stats = {"procesados": 0, "exitosos": 0, "fallidos": 0}
        
        excel_app = None
        wb = None
        
        try:
            # 1. Leer datos
            rows, keys = self.read_data_sheet(
                config['archivo_excel'],
                config.get('hoja_datos', 'DATA'),
                config.get('columna_clave', 17)
            )
            
            if not rows:
                log.warning("No se encontraron datos")
                return stats
            
            total = len(rows)
            
            # 2. Generar QRs en batch
            temp_folder = Path(os.environ['TEMP']) / f"qr_batch_{int(time.time())}"
            qr_paths = self.generate_qr_batch(keys, temp_folder)
            
            # 3. Iniciar Excel COM
            log.info("Iniciando Excel...")
            pythoncom.CoInitialize()
            
            excel_app = win32com.client.DispatchEx("Excel.Application")
            excel_app.Visible = False
            excel_app.DisplayAlerts = False
            
            # Abrir archivo
            ruta_abs = os.path.abspath(config['archivo_excel'])
            wb = excel_app.Workbooks.Open(ruta_abs)
            ws_plantilla = wb.Sheets(config['hoja_plantilla'])
            
            # 4. Procesar cada registro
            for idx, row in enumerate(rows):
                if self.stop_requested:
                    log.warning("Detenido por usuario")
                    break
                
                key = keys[idx]
                
                if progress_callback:
                    progress_callback(idx + 1, total, key)
                
                try:
                    # Escribir datos en plantilla
                    start_col = 18  # Columna R
                    for c_idx, valor in enumerate(row):
                        ws_plantilla.Cells(4, start_col + c_idx).Value = valor
                    
                    # Insertar QR
                    try:
                        ws_plantilla.Shapes("CODIGO_QR_ACTUAL").Delete()
                    except:
                        pass
                    
                    qr_path = qr_paths.get(idx)
                    if qr_path and os.path.exists(qr_path):
                        celda_qr = ws_plantilla.Range("M8")
                        pic = ws_plantilla.Shapes.AddPicture(
                            qr_path,
                            False, True,
                            celda_qr.Left, celda_qr.Top,
                            60, 60
                        )
                        pic.Name = "CODIGO_QR_ACTUAL"
                    
                    # Exportar PDF
                    if config.get('generar_pdf', True):
                        nombre_pdf = "".join([c for c in str(key) if c.isalnum() or c in " .-_"])
                        ruta_pdf = os.path.join(config['carpeta_pdf'], f"{nombre_pdf}.pdf")
                        ws_plantilla.ExportAsFixedFormat(0, ruta_pdf)
                        log.debug(f"PDF: {nombre_pdf}.pdf")
                    
                    # Imprimir
                    if config.get('imprimir', False):
                        impresora = config.get('impresora')
                        copias = config.get('copias', 1)
                        
                        try:
                            if impresora:
                                ws_plantilla.PrintOut(Copies=copias, ActivePrinter=impresora)
                            else:
                                ws_plantilla.PrintOut(Copies=copias)
                        except Exception as e:
                            log.warning(f"Error imprimiendo: {e}")
                    
                    stats["procesados"] += 1
                    stats["exitosos"] += 1
                    
                except Exception as e:
                    log.error(f"Error en registro {idx+1}: {e}")
                    stats["fallidos"] += 1
            
        except Exception as e:
            log.error(f"Error crítico: {e}")
            
        finally:
            # Cerrar Excel
            if wb:
                wb.Close(SaveChanges=False)
            if excel_app:
                excel_app.Quit()
            pythoncom.CoUninitialize()
            
            # Registrar operación
            log_operation(
                module="excel_processor",
                action="process_excel",
                success=stats['fallidos'] == 0,
                message=f"Procesados {stats['exitosos']} de {stats['procesados']}",
                total=stats['procesados'],
                success_count=stats['exitosos'],
                error_count=stats['fallidos']
            )
            
            log.info("Proceso Excel finalizado")
            
        return stats
    
    def process_simple(self,
                      excel_path: str | Path,
                      output_folder: str | Path,
                      key_column: int = 0,
                      progress_callback: Optional[Callable] = None) -> Dict[str, int]:
        """
        Procesamiento simple: lee Excel y genera QRs para cada fila.
        
        Modo simplificado sin necesidad de Excel COM.
        
        Args:
            excel_path: Ruta al Excel
            output_folder: Carpeta de salida para QRs
            key_column: Columna clave (0-indexed)
            progress_callback: Función(idx, total, key)
            
        Returns:
            Estadísticas
        """
        self.stop_requested = False
        stats = {"procesados": 0, "exitosos": 0, "fallidos": 0}
        
        try:
            output_folder = Path(output_folder)
            output_folder.mkdir(parents=True, exist_ok=True)
            
            # Leer Excel
            log.info(f"Leyendo {excel_path}...")
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            ws = wb.active
            
            # Contar filas
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            total = len([r for r in rows if r[0] is not None])
            
            log.info(f"Procesando {total} registros...")
            
            idx = 0
            for row in rows:
                if self.stop_requested:
                    break
                    
                if row[0] is None:
                    continue
                
                idx += 1
                key = str(row[key_column]) if len(row) > key_column and row[key_column] else f"item_{idx}"
                
                if progress_callback:
                    progress_callback(idx, total, key)
                
                try:
                    # Generar QR
                    qr_path = output_folder / f"{key}.png"
                    success, _ = self.qr_generator.generate_qr(
                        content=key,
                        output_path=str(qr_path),
                        size=300
                    )
                    
                    if success:
                        stats["exitosos"] += 1
                    else:
                        stats["fallidos"] += 1
                    
                    stats["procesados"] += 1
                    
                except Exception as e:
                    log.error(f"Error en fila {idx}: {e}")
                    stats["fallidos"] += 1
            
            wb.close()
            
            # Registrar
            log_operation(
                module="excel_processor",
                action="process_simple",
                success=stats['fallidos'] == 0,
                message=f"Generados {stats['exitosos']} QRs de {total} registros"
            )
            
        except Exception as e:
            log.error(f"Error: {e}")
            
        return stats
